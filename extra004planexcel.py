import openpyxl
# importando biblioteca

book = openpyxl.Workbook()
# criar uma planilha book
print(book.sheetnames)
# lendo paginas existentes
book.create_sheet('Frutas')
# criando nova aba na planilha
frutas_page = book['Frutas']
# selecionando uma aba na planilha
frutas_page.append(['Nome da fruta', 'Quantidade', 'Valor'])
# cria o título das colunas
frutas_page.append(['Banana','5','R$ 3,90'])
frutas_page.append(['Mamão','10','R$ 4,50'])
frutas_page.append(['Uva','7','R$ 5,50'])
# tantas repetições quantos dados forém ser inseridos
book.save('Planilhe de Teste.xlsx')
# salvando a modificação na planilha
